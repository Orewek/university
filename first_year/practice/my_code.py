import csv

from get_median_frame import make_frame

from hist_median_frame import make_hist


from openpyxl import load_workbook


def get_data(list_min, list_hour, list_folders, mega_frame, bad_videos):
    with open('data.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)

        for i in range(len(list_hour)):
            for j in range(len(list_min) - 1):
                time: str = f'{list_hour[i]}.{list_min[j]}'
                print(time)
                if (f'{list_hour[i]}.{list_min[j]}' in bad_videos):
                    continue

                path = f'{list_hour[i]}.{list_min[j]}.00-{list_hour[i]}.{list_min[j + 1]}.00[M][0@0][0]'
                if (list_min[j] == '55'):
                    path = f'{list_hour[i]}.{list_min[j]}.00-{list_hour[i + 1]}.{list_min[j + 1]}.00[M][0@0][0]'

                frame_folder: str = list_folders[i]
                video_folder: str = f'video/archive{list_hour[i]}'
                make_frame(path, frame_folder, mega_frame, video_folder)
                data = make_hist(path, frame_folder, mega_frame)

                writer.writerow(data)


def input_to_excel(excel_file, csv_file):
    workbook = load_workbook(filename=excel_file)
    ws4 = workbook["Sheet1"]

    start_row = 148
    with open(csv_file) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        next(csv_reader, None)

        for row in csv_reader:
            available_status = ws4.cell(row=start_row, column=10).value
            while (available_status is not None):
                start_row += 1
                available_status = ws4.cell(row=start_row, column=10).value

            ws4.cell(row=start_row, column=12).value: str = row[0]
            ws4.cell(row=start_row, column=13).value: str = row[1]
            ws4.cell(row=start_row, column=14).value: str = row[2]
            ws4.cell(row=start_row, column=15).value: str = f'{row[3]}; {row[4]}'
            ws4.cell(row=start_row, column=16).value: str = f'{row[5]}; {row[6]}'

            start_row += 1

    workbook.save(excel_file)


if __name__ == '__main__':
    list_min: list = ['00', '05', '10', '15', '20', '25', '30',
                      '35', '40', '45', '50', '55', '00']
    # list_folders: list = ['8.00-9.00 07.07.2022',
    #                 '9.00-10.00 07.07.2022',
    #                 '10.00-11.00 07.07.2022',
    #                 '11.00-12.00 07.07.2022']
    # mega_frame: list = '8.00-12.00 07.07.2022'
    # bad_videos: list = ['08.05', '08.35', '08.40', '09.35', '09.40',
    #               '09.45', '09.55', '10.40', '10.45', '11.55']
    # list_hour: list = ['08', '09', '10', '11']

    # list_folders: list = ['22.00-23.00 11.06.2023',
    #                 '23.00-24.00 11.06.2023',
    #                 '00.00-01.00 12.06.2023',
    #                 '01.00-02.00 12.06.2023']
    # mega_frame: list = '22.00-02.00 11-12.06.2023'
    # bad_videos: list = ['22.30', '23.10', '00.00', '00.25',
    #               '00.50', '01.50', '01.55']
    # list_hour: list = ['22', '23', '00', '01', '02']

    # list_folders: list = ['22.00-23.00 12.06.2023',
    #                 '23.00-24.00 12.06.2023',
    #                 '00.00-01.00 13.06.2023',
    #                 '01.00-02.00 13.06.2023']
    # mega_frame: list = '22.00-02.00 12-13.06.2023'
    # bad_videos: list = ['22.45', '23.45', '00.05', '00.35',
    #               '00.50', '01.10', '01.15',
    #               '01.30', '01.40', '01.45', '01.50']
    # list_hour: list = ['22', '23', '00', '01', '02']

    list_folders: list = ['22.00-23.00 13.06.2023',
                          '23.00-24.00 13.06.2023',
                          '00.00-01.00 14.06.2023',
                          '01.00-02.00 14.06.2023']
    mega_frame: list = '22.00-02.00 13-14.06.2023'
    bad_videos: list = ['22.30', '22.50', '23.20',
                        '23.25', '23.35', '00.15',
                        '01.10', '01.15', '01.20']
    list_hour: list = ['22', '23', '00', '01']

    header: list = ['math expect, disperc, sqrt mean, max_x, max_y, min_x, min_y']
    # get_data(list_min, list_hour, list_folders, mega_frame, bad_videos)

    excel_file: str = 'Разметка видео.xlsx'
    input_to_excel(excel_file, 'data.csv')
